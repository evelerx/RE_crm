from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from ..auth import get_current_user
from ..enterprise_scope import assign_enterprise_fields, user_can_access_record, user_read_filter
from ..models import Activity, Contact, Deal, User
from ..db import get_session


router = APIRouter(prefix="/csv", tags=["csv"])


CONTACT_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("name", "client name", "contact name", "customer name", "full name", "lead name"),
    "occupation": ("occupation", "profession", "job", "designation", "work"),
    "role": ("role", "purpose", "client type", "type", "category"),
    "phone": ("phone", "mobile", "mobile number", "phone number", "contact number", "contact no"),
    "email": ("email", "email id", "mail", "email address", "client email"),
    "tags": ("tags", "tag", "feedback", "segment", "labels"),
    "notes": ("notes", "remarks", "comment", "comments", "description"),
}

DEAL_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "title": ("title", "property name", "deal name", "listing name", "project name", "property"),
    "asset_type": ("asset type", "property type", "asset", "type"),
    "stage": ("stage", "deal stage", "status"),
    "city": ("city",),
    "area": ("area", "location", "locality", "micro market", "submarket"),
    "visit_date": ("visit date", "date of visit", "site visit date", "visit on"),
    "typology": ("typology", "configuration", "requirement", "client requirement", "bhk", "unit type"),
    "ticket_size": ("ticket size", "asking price", "budget", "price", "deal value", "amount"),
    "customer_budget": ("customer budget", "client budget", "budget amount", "buyer budget"),
    "expected_yield_pct": ("expected yield pct", "yield", "yield (%)", "rental yield", "rental yield (%)"),
    "expected_roi_pct": ("expected roi pct", "roi", "roi (%)", "expected roi", "expected roi (%)"),
    "liquidity_days_est": ("liquidity days est", "liquidity", "liquidity days", "days to liquidate"),
    "client_phase": ("client phase", "phase", "temperature", "lead temperature", "client heat"),
    "close_probability": ("close probability", "close probability (%)", "probability", "win probability"),
    "risk_flags": ("risk flags", "risk", "risks", "risk score", "risk score (1-10)"),
    "contact_email": ("contact email", "client email", "email", "email id"),
    "notes": ("notes", "remarks", "comment", "comments", "description", "last interaction", "next follow-up"),
    "contact_name": ("client name", "contact name", "customer name", "lead name"),
}


def _csv_response(filename: str, content: str) -> StreamingResponse:
    bio = io.BytesIO(content.encode("utf-8-sig"))
    return StreamingResponse(
        bio,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _normalize_header(value: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else " " for ch in (value or "")).strip()


def _load_tabular_rows(file: UploadFile) -> list[dict[str, str]]:
    filename = (file.filename or "").lower()
    raw = file.file.read()

    if filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise HTTPException(
                status_code=500,
                detail="Excel import requires openpyxl. Install backend requirements or use CSV for now.",
            ) from exc
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        out: list[dict[str, str]] = []
        for values in rows[1:]:
            row: dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = values[idx] if idx < len(values) else None
                row[header] = "" if value is None else str(value).strip()
            out.append(row)
        return out

    text = raw.decode("utf-8-sig", errors="ignore")
    return [{k: (v or "").strip() for k, v in row.items() if k} for row in csv.DictReader(io.StringIO(text))]


def _match_value(row: dict[str, str], aliases: tuple[str, ...]) -> str:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_normalize_header(alias), "")
        if value:
            return value.strip()
    return ""


def _deal_title_from_row(row: dict[str, str]) -> str:
    title = _match_value(row, DEAL_FIELD_ALIASES["title"])
    if title:
        return title
    contact_name = _match_value(row, DEAL_FIELD_ALIASES["contact_name"])
    area = _match_value(row, ("location", "area", "city"))
    if contact_name and area:
        return f"{contact_name} - {area}"
    return contact_name or area


def _float_value(value: str) -> Optional[float]:
    if not value:
        return None
    cleaned = (
        value.replace(",", "")
        .replace("₹", "")
        .replace("%", "")
        .replace("rs.", "")
        .replace("rs", "")
        .strip()
    )
    try:
        return float(cleaned)
    except Exception:
        return None


def _int_value(value: str) -> Optional[int]:
    num = _float_value(value)
    if num is None:
        return None
    try:
        return int(round(num))
    except Exception:
        return None


def _date_value(value: str) -> Optional[date]:
    raw = (value or "").strip()
    if not raw:
        return None
    normalized = raw.replace(".", "-").replace("/", "-")
    for parser in (date.fromisoformat,):
        try:
            return parser(normalized)
        except Exception:
            pass
    for fmt in ("%d-%m-%Y", "%m-%d-%Y", "%d-%m-%y", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(normalized, fmt).date()
        except Exception:
            continue
    return None


def _normalize_asset_type(value: str) -> str:
    normalized = _normalize_header(value)
    mapping = {
        "residential": "residential",
        "commercial": "commercial",
        "land": "land",
        "industrial": "industrial",
        "other": "other",
    }
    return mapping.get(normalized, "residential")


def _normalize_stage(value: str) -> str:
    normalized = _normalize_header(value)
    mapping = {
        "lead": "lead",
        "visit": "visit",
        "site visit": "visit",
        "negotiation": "negotiation",
        "closed": "closed",
        "won": "closed",
        "lost": "lost",
        "dropped": "lost",
    }
    return mapping.get(normalized, "lead")


def _normalize_client_phase(value: str) -> str:
    normalized = _normalize_header(value)
    mapping = {
        "hot": "hot",
        "warm": "warm",
        "cold": "cold",
        "lost": "lost",
    }
    return mapping.get(normalized, "")


@router.get("/export/contacts")
def export_contacts(
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    rows = session.exec(
        select(Contact).where(user_read_filter(Contact, user)).order_by(Contact.created_at.desc())
    ).all()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["name", "occupation", "purpose", "phone", "email", "feedback", "notes"])
    for c in rows:
        w.writerow([c.name, c.occupation or "", c.role, c.phone or "", c.email or "", c.tags or "", c.notes or ""])
    return _csv_response("contacts.csv", out.getvalue())


@router.get("/export/deals")
def export_deals(
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    rows = session.exec(
        select(Deal).where(user_read_filter(Deal, user)).order_by(Deal.created_at.desc())
    ).all()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(
        [
            "title",
            "asset_type",
            "stage",
            "city",
            "location",
            "visit_date",
            "typology",
            "ticket_size",
            "customer_budget",
            "client_phase",
            "expected_yield_pct",
            "expected_roi_pct",
            "liquidity_days_est",
            "close_probability",
            "risk_flags",
            "contact_email",
            "notes",
        ]
    )
    for d in rows:
        contact_email = ""
        if d.contact_id:
            c = session.get(Contact, d.contact_id)
            if c and user_can_access_record(c, user) and c.email:
                contact_email = c.email
        w.writerow(
            [
                d.title,
                d.asset_type,
                d.stage,
                d.city,
                d.area,
                d.visit_date.isoformat() if d.visit_date else "",
                d.typology or "",
                d.ticket_size if d.ticket_size is not None else "",
                d.customer_budget if d.customer_budget is not None else "",
                d.client_phase or "",
                d.expected_yield_pct if d.expected_yield_pct is not None else "",
                d.expected_roi_pct if d.expected_roi_pct is not None else "",
                d.liquidity_days_est if d.liquidity_days_est is not None else "",
                d.close_probability if d.close_probability is not None else "",
                d.risk_flags or "",
                contact_email,
                d.notes or "",
            ]
        )
    return _csv_response("deals.csv", out.getvalue())


@router.get("/export/activities")
def export_activities(
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    rows = session.exec(
        select(Activity).where(user_read_filter(Activity, user)).order_by(Activity.created_at.desc())
    ).all()

    deal_ids = {a.deal_id for a in rows if a.deal_id}
    contact_ids = {a.contact_id for a in rows if a.contact_id}

    deal_title: dict[UUID, str] = {}
    contact_name: dict[UUID, str] = {}

    if deal_ids:
        for d in session.exec(select(Deal).where(Deal.id.in_(list(deal_ids)))).all():
            if user_can_access_record(d, user):
                deal_title[d.id] = d.title
    if contact_ids:
        for c in session.exec(select(Contact).where(Contact.id.in_(list(contact_ids)))).all():
            if user_can_access_record(c, user):
                contact_name[c.id] = c.name

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["kind", "summary", "due_at", "completed", "deal_title", "contact_name", "created_at"])
    for a in rows:
        w.writerow(
            [
                a.kind,
                a.summary or "",
                a.due_at.isoformat() + "Z" if a.due_at else "",
                "1" if a.completed else "0",
                deal_title.get(a.deal_id, "") if a.deal_id else "",
                contact_name.get(a.contact_id, "") if a.contact_id else "",
                a.created_at.isoformat() + "Z",
            ]
        )
    return _csv_response("activities.csv", out.getvalue())


@router.post("/import/contacts")
async def import_contacts(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    await file.seek(0)
    rows = _load_tabular_rows(file)
    created = 0
    for row in rows:
        name = _match_value(row, CONTACT_FIELD_ALIASES["name"])
        if not name:
            continue
        contact = Contact(
            name=name,
            occupation=_match_value(row, CONTACT_FIELD_ALIASES["occupation"]),
            role=_match_value(row, CONTACT_FIELD_ALIASES["role"]) or "buyer",
            phone=_match_value(row, CONTACT_FIELD_ALIASES["phone"]) or None,
            email=_match_value(row, CONTACT_FIELD_ALIASES["email"]) or None,
            tags=_match_value(row, CONTACT_FIELD_ALIASES["tags"]),
            notes=_match_value(row, CONTACT_FIELD_ALIASES["notes"]),
        )
        assign_enterprise_fields(contact, user)
        session.add(contact)
        created += 1
    session.commit()
    return {"created": created}


@router.post("/import/deals")
async def import_deals(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    user: User = Depends(get_current_user),
):
    await file.seek(0)
    rows = _load_tabular_rows(file)
    created = 0

    # Map contact emails to ids for linking
    contacts = session.exec(select(Contact).where(user_read_filter(Contact, user))).all()
    email_to_id = {c.email.lower(): c.id for c in contacts if c.email}

    for row in rows:
        title = _deal_title_from_row(row).strip()
        if not title:
            continue
        contact_id: Optional[UUID] = None
        contact_email = _match_value(row, DEAL_FIELD_ALIASES["contact_email"]).lower()
        if contact_email and contact_email in email_to_id:
            contact_id = email_to_id[contact_email]

        deal = Deal(
            title=title,
            asset_type=_normalize_asset_type(_match_value(row, DEAL_FIELD_ALIASES["asset_type"])),
            stage=_normalize_stage(_match_value(row, DEAL_FIELD_ALIASES["stage"])),
            city=_match_value(row, DEAL_FIELD_ALIASES["city"]),
            area=_match_value(row, DEAL_FIELD_ALIASES["area"]),
            visit_date=_date_value(_match_value(row, DEAL_FIELD_ALIASES["visit_date"])),
            typology=_match_value(row, DEAL_FIELD_ALIASES["typology"]),
            ticket_size=_float_value(_match_value(row, DEAL_FIELD_ALIASES["ticket_size"])),
            customer_budget=_float_value(_match_value(row, DEAL_FIELD_ALIASES["customer_budget"])),
            expected_yield_pct=_float_value(_match_value(row, DEAL_FIELD_ALIASES["expected_yield_pct"])),
            expected_roi_pct=_float_value(_match_value(row, DEAL_FIELD_ALIASES["expected_roi_pct"])),
            liquidity_days_est=_int_value(_match_value(row, DEAL_FIELD_ALIASES["liquidity_days_est"])),
            client_phase=_normalize_client_phase(_match_value(row, DEAL_FIELD_ALIASES["client_phase"])),
            close_probability=_int_value(_match_value(row, DEAL_FIELD_ALIASES["close_probability"])),
            risk_flags=_match_value(row, DEAL_FIELD_ALIASES["risk_flags"]),
            contact_id=contact_id,
            notes=_match_value(row, DEAL_FIELD_ALIASES["notes"]),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        assign_enterprise_fields(deal, user)
        session.add(deal)
        created += 1

    session.commit()
    return {"created": created}
